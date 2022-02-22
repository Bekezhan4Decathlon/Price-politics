# import pandas as pd
import random
import time
from tqdm import tqdm
import hmac
import hashlib
import requests
import os

import openpyxl
from openpyxl.drawing.image import Image

from COP_parser.cop_parse import Parser
from env import key, image_codes

data_path = os.path.dirname(os.path.realpath(__file__))

HEADERS = [
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36 OPR/68.0.3618.125',
        'accept': '*/*',
    },
    {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.3 Safari/605.1.15',
        'accept': '*/*'
    },
    {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 7.0; SM-G930VC Build/NRD90M; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/58.0.3029.83 Mobile Safari/537.36',
        'accept': '*/*'
    },
    {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1',
        'accept': '*/*'
    }
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36 OPR/68.0.3618.125",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.3 Safari/605.1.15",
    "Mozilla/5.0 (Linux; Android 7.0; SM-G930VC Build/NRD90M; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/58.0.3029.83 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1"
]

parser = Parser(USER_AGENTS)

def download_image(model_code):
    if os.path.exists(f'{data_path}/images') == False:
        os.mkdir(f"{data_path}/images")

    image_code = image_codes[str(model_code)]
    hash = hmac.new(bytes(key, 'utf-8'), bytes(f"p{int(image_code)}", 'utf-8'), hashlib.md5).hexdigest()
    url = f"https://contents.mediadecathlon.com/p{int(image_code)}/k${hash}/image.jpg?format=auto&quality=75&f=80x80"
    response = requests.get(url)

    file = open(f"{data_path}/images/{model_code}.png", "wb")
    file.write(response.content)
    file.close()

    return f"{data_path}/images/{model_code}.png"

if __name__ == '__main__':
    excel_file = openpyxl.load_workbook('COP 0.xlsx')
    for sheet in tqdm(excel_file.sheetnames):
        COP_sheet = excel_file[sheet]
        for c in range(1, COP_sheet.max_column+1):
            try:
                if COP_sheet.cell(row=1, column=c).value.lower() == "sportmaster":
                    for r in range(2, COP_sheet.max_row+1):
                        # image_filename_path = download_image(COP_sheet.cell(row=r, column=1).value)
                        COP_sheet.cell(row=r, column=c).hyperlink = COP_sheet.cell(row=r, column=c).value
                        product_price = parser.get_price_sportmaster(COP_sheet.cell(row=r, column=c).value)
                        COP_sheet.cell(row=r, column=c).value = product_price
                        # COP_sheet.add_image(Image(image_filename_path), f'C{r}')
                        time.sleep(random.randrange(4, 8))

                elif COP_sheet.cell(row=1, column=c).value.lower() == "prosport/limpopo":
                    for r in range(2, COP_sheet.max_row+1):
                        COP_sheet.cell(row=r, column=c).hyperlink = COP_sheet.cell(row=r, column=c).value
                        if "limpopo" in COP_sheet.cell(row=r, column=c).value.lower():
                            product_price = parser.get_price_limpopo(COP_sheet.cell(row=r, column=c).value)
                        elif "prosport" in COP_sheet.cell(row=r, column=c).value.lower():
                            product_price = parser.get_price_prosport(COP_sheet.cell(row=r, column=c).value)
                        COP_sheet.cell(row=r, column=c).value = product_price
                        time.sleep(random.randrange(4, 8))

                elif COP_sheet.cell(row=1, column=c).value.lower() == "kaspi":
                    for r in range(2, COP_sheet.max_row+1):
                        COP_sheet.cell(row=r, column=c).hyperlink = COP_sheet.cell(row=r, column=c).value
                        product_price = parser.get_price_kaspi(COP_sheet.cell(row=r, column=c).value)
                        COP_sheet.cell(row=r, column=c).value = product_price
                        time.sleep(random.randrange(4, 8))
            except Exception as e:
                print(e)

    excel_file.save('COP 0_output.xlsx')
